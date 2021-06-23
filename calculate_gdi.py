# -*- coding: utf-8 -*-
"""

Calculate GDI for multiple c3d files.

@author: Jussi (jnu@iki.fi)
"""


# %% init
import pathlib
from time import time
import gaitutils
from gaitutils import nexus
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.styles import Font
import csv


def _bold_cell(ws, **cell_params):
    """Write a bold-styled cell into worksheet ws"""
    boldfont = Font(bold=True)
    _cell = ws.cell(**cell_params)
    _cell.font = boldfont


def _auto_adjust(ws):
    """Auto adjust column widths of worksheet ws"""
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


def _compute_gdi(c3dfile):
    """Compute GDI in Nexus for a given c3d file"""
    gdi_pipeline = 'Calculate GDI'  # Nexus pipeline name
    gaitutils.nexus._open_trial(c3dfile)
    gaitutils.nexus._run_pipelines(gdi_pipeline)


def _read_gdi():
    """Read GDI values from current Nexus trial"""
    gdi = {'Left': None, 'Right': None}
    subj = gaitutils.nexus.get_subjectnames()
    for context in gdi.keys():
        this_gdi, exists = vicon.GetAnalysisParam(subj, f'{context}GDI')
        if not exists:
            print('GDI value not available in Nexus')
            # raise RuntimeError('GDI value not available in Nexus')
        else:
            gdi[context] = this_gdi
    return gdi


vicon = gaitutils.nexus.viconnexus()


# %% some test files
testdir = r"C:\Temp\D0061_VH\2019_2_11_PostOp10kk_VH"
c3dfiles = gaitutils.sessionutils.get_c3ds(testdir, trial_type='dynamic')


# %% the actual files from CSV
csvfn = r"Y:\Userdata_Vicon_Server\Testing\collect_gait_testit\leikattujen_tiedostopolut2.csv"
exclude_list = ['2015_1_15_preop_IS']  # excluded due to missing subject info etc.
drive_letter = 'Y:'
c3dfiles = list()
# the csv seems to use utf-8 encoding instead of Windows default (cp1252)
types = set()
with open(csvfn, encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile)
    SKIP_ROWS = 1  # skip this many rows in the beginning
    for k, row in enumerate(reader):
        if k < SKIP_ROWS:
            continue
        if len(row) > 1:
            raise RuntimeError('rows should only contain one element (file path)')
        row = row[0]
        # this is only needed for paths beginning with \\husnasdb
        # row = row[row.find('Vicon_data') + 10 :]
        # path = pathlib.Path(drive_letter + row)
        path = pathlib.Path(row)
        if any(excl in str(path) for excl in exclude_list):
            print(f'in exclude list: {path}')
            continue
        if not path.exists():
            raise FileNotFoundError(path)
        # get the corresponding ENF file and check if trial is dynamic
        enffile = pathlib.Path(str(path).replace('.c3d', '.Trial.enf'))
        if not enffile.exists():
            # try alternative naming
            ind = str(path).find('.c3d') - 2
            nn = str(path)[ind : ind + 2]
            enffile = pathlib.Path(str(path).replace('.c3d', f'.Trial{nn}.enf'))
            if not enffile.exists():  # give up
                raise FileNotFoundError(enffile)
        edata = gaitutils.eclipse.get_eclipse_keys(enffile, return_empty=True)
        if 'TYPE' not in edata:
            raise RuntimeError('no type specified in ENF')
        types.add(edata['DESCRIPTION'])
        if 'dynamic' in edata['TYPE'].lower():
            c3dfiles.append(path)
        else:
            print(f'{path} is seemingly not a dynamic trial, excluding')
        if k % 10 == 0:
            print(f'{k} files analyzed')
print(f'{len(c3dfiles)} valid c3d files found')


# %% DEBUG: let's get trials with unique sessions
c3dfiles_unique = list()
sessions_seen = set()
for c3dfile in c3dfiles:
    session = c3dfile.parent
    if session not in sessions_seen:
        sessions_seen.add(session)
        c3dfiles_unique.append(c3dfile)


# %% DEBUG: visit all sessions quickly
# this is to make sure Nexus opens them without issues
import time
import os.path as op

for k, c3dfile in enumerate(list(reversed(c3dfiles_unique))[80:], 1):
    if any(excl in str(c3dfile) for excl in exclude_list):
        print(f'in exclude list: {path}')
        continue
    print(f'opening {c3dfile}')
    trialpath_ = op.splitext(c3dfile)[0]
    vicon.OpenTrial(trialpath_, 60)
    assert len(vicon.GetSubjectNames()) == 1


# %% the real deal: compute GDIs for all trials, write into XLSX
# define the 'c3dfiles' variable first

fname_xls = rf"C:\Temp\gdi_{pathlib.Path(csvfn).stem}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
_bold_cell(ws, column=2, row=5, value='GDI, oik.')
_bold_cell(ws, column=2, row=6, value='GDI, vas.')

counter = 1
for col, c3dfile in enumerate(c3dfiles, 3):
    _compute_gdi(c3dfile)
    vicon.SaveTrial(60)
    try:
        gdi = _read_gdi()
    except gaitutils.GaitDataError:
        print(
            f'***skipping trial {c3dfile}, multiple subjects in session or other error'
        )
        continue
    _bold_cell(ws, column=col, row=4, value=str(c3dfile))
    ws.cell(column=col, row=5, value=gdi['Right'])
    ws.cell(column=col, row=6, value=gdi['Left'])
    print(f'{c3dfile}: {gdi}')
    if counter % 10 == 0:
        print(f'{counter}/{len(c3dfiles)} files processed')
    counter += 1

# _auto_adjust(ws)
wb.save(filename=fname_xls)
